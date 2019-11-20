import openpyxl as xl

wb = xl.load_workbook("test.xlsx")
# print(wb)
wb.create_sheet("新建页面", index=2)

listSheet = wb.sheetnames  # 加载所有的sheet页
sheet = wb[listSheet[0]]  # 选择一个sheet页
sheet2 = wb[listSheet[2]]

# 复制sheet
# for i in range(1, sheet.max_row + 1):
#     sheet2['A{0}'.format(i)].value = sheet['A{0}'.format(i)].value
#     sheet2['B{0}'.format(i)].value = sheet['B{0}'.format(i)].value
for i in range(1, sheet2.max_row+1):
    sheet2['C{0}'.format(i)] = '=SUBSTITUTE(A{0},"哈登",B{1})'.format(i, i)

wb.save("jieguo5.xlsx")